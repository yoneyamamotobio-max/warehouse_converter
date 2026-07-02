from __future__ import annotations

import ctypes
import json
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QRect, Qt
from PySide6.QtGui import QColor, QDragEnterEvent, QDropEvent, QFont, QIcon, QPainter, QPen
from PySide6.QtWidgets import (
    QApplication,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QLabel,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QStyle,
    QVBoxLayout,
    QWidget,
)


APP_ORG = "WarehouseTools"
APP_NAME = "InventoryExcelConverter"
APP_USER_MODEL_ID = "WarehouseTools.InventoryExcelConverter"
GRID_COLUMNS = 12
GRID_ROWS = 23
AISLE_COLUMN_LABELS = {"B", "E", "F", "J"}

DETAIL_COLUMNS = [
    "品番",
    "サイズ",
    "厚み",
    "加工/裏表",
    "グレード",
    "枚数",
    "ロット",
    "備考",
    "入庫日",
    "在庫日数",
    "パレット番号",
    "保管場所",
    "色",
    "更新日時",
]
SUMMARY_COLUMNS = ["品番", "サイズ", "厚み", "加工/裏表", "グレード", "ロット", "合計枚数"]
PALLET_COLUMNS = ["パレット番号", "保管場所", "入庫日", "在庫日数", "色", "更新日時", "明細数", "合計枚数"]
MAJOR_COLUMN_PRIORITY = [
    "品番",
    "part_code",
    "items.part_code",
    "サイズ",
    "size",
    "items.size",
    "厚み",
    "thickness_mm",
    "items.thickness_mm",
    "加工/裏表",
    "finish_text",
    "items.finish_text",
    "グレード",
    "grade",
    "items.grade",
    "枚数",
    "sheet_count",
    "items.sheet_count",
    "ロット",
    "lot",
    "items.lot",
    "備考",
    "note",
    "items.note",
    "入庫日",
    "received_date",
    "在庫日数",
    "パレット番号",
    "pallet_number",
    "保管場所",
    "location_code",
    "色",
    "color_key",
    "更新日時",
    "updated_at",
]
NUMERIC_COLUMNS = {"枚数", "合計枚数", "明細数", "在庫日数", "sheet_count", "items.sheet_count"}
TEXT_COLUMNS = {"パレット番号", "pallet_number"}


@dataclass
class TableData:
    name: str
    columns: list[str]
    rows: list[dict[str, Any]]


def resource_path(filename: str) -> Path:
    base_path = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return base_path / filename


def set_windows_app_user_model_id() -> None:
    if sys.platform != "win32":
        return
    try:
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(APP_USER_MODEL_ID)
    except Exception:
        pass


def app_icon() -> QIcon:
    icon_path = resource_path("icon.ico")
    return QIcon(str(icon_path)) if icon_path.exists() else QIcon()


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return str(value)


def as_int(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def column_label(index: int) -> str:
    index += 1
    text = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        text = chr(65 + remainder) + text
    return text


def is_aisle_column(index: int) -> bool:
    return column_label(index) in AISLE_COLUMN_LABELS


def grid_column_display_label(index: int) -> str:
    label = column_label(index)
    return f"{label}(通路)" if label in AISLE_COLUMN_LABELS else label


def is_blocked_grid_cell(col: int, row: int) -> bool:
    label = column_label(col)
    number = row + 1
    if label in {"A", "B"}:
        return 20 <= number <= 23
    if label == "C":
        return 19 <= number <= 23
    if label in {"D", "H"}:
        return 16 <= number <= 23
    if label in {"E", "F"}:
        return 17 <= number <= 23
    if label in {"G", "I"}:
        return 18 <= number <= 23
    if label == "J":
        return 22 <= number <= 23
    if label == "K":
        return number in {2, 3, 5, 6, 8, 9, 10, 12, 13, 15, 16, 17, 18, 20, 21, 22, 23}
    if label == "L":
        return 1 <= number <= 23
    return False


def inventory_days(received_date: Any, today: date) -> int | None:
    text = as_text(received_date).strip()
    if not text:
        return None
    try:
        parsed = datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        return None
    return (today - parsed).days


def first_value(source: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in source:
            return source.get(key)
    return ""


def flatten_value(value: Any, prefix: str = "") -> dict[str, Any]:
    if isinstance(value, dict):
        result: dict[str, Any] = {}
        for key, child in value.items():
            name = f"{prefix}.{key}" if prefix else str(key)
            result.update(flatten_value(child, name))
        return result
    if isinstance(value, list):
        return {prefix: json.dumps(value, ensure_ascii=False, separators=(",", ":"))}
    return {prefix: value}


def ordered_columns(rows: list[dict[str, Any]], preferred: list[str] | None = None) -> list[str]:
    seen: set[str] = set()
    columns: list[str] = []
    present = {column for row in rows for column in row}
    for column in preferred or []:
        if column in present and column not in seen:
            seen.add(column)
            columns.append(column)
    for row in rows:
        for column in row:
            if column not in seen:
                seen.add(column)
                columns.append(column)
    return columns


def build_inventory_detail(pallets: list[Any], today: date) -> TableData:
    rows: list[dict[str, Any]] = []
    for pallet in pallets:
        if not isinstance(pallet, dict):
            continue
        items = pallet.get("items")
        if not isinstance(items, list) or not items:
            continue
        for item in items:
            if not isinstance(item, dict):
                item = {}
            received_date = first_value(pallet, "received_date", "received_at")
            rows.append(
                {
                    "品番": as_text(first_value(item, "part_code", "product_code", "品番")),
                    "サイズ": as_text(first_value(item, "size", "サイズ")),
                    "厚み": as_text(first_value(item, "thickness_mm", "thickness", "厚み")),
                    "加工/裏表": as_text(first_value(item, "finish_text", "finish", "加工/裏表")),
                    "グレード": as_text(first_value(item, "grade", "グレード")),
                    "枚数": as_int(first_value(item, "sheet_count", "count", "枚数")),
                    "ロット": as_text(first_value(item, "lot", "ロット")),
                    "備考": as_text(first_value(item, "note", "備考")),
                    "入庫日": as_text(received_date),
                    "在庫日数": inventory_days(received_date, today),
                    "パレット番号": as_text(first_value(pallet, "pallet_number", "pallet_no")),
                    "保管場所": as_text(first_value(pallet, "location_code", "location")),
                    "色": as_text(first_value(pallet, "color_key", "color", "color_mode")),
                    "更新日時": as_text(first_value(pallet, "updated_at")),
                }
            )
    return TableData("在庫明細", DETAIL_COLUMNS, rows)


def build_inventory_summary(detail: TableData) -> TableData:
    totals: dict[tuple[str, ...], int] = defaultdict(int)
    key_columns = ["品番", "サイズ", "厚み", "加工/裏表", "グレード", "ロット"]
    for row in detail.rows:
        key = tuple(as_text(row.get(column)) for column in key_columns)
        totals[key] += as_int(row.get("枚数"))

    rows = []
    for key in sorted(totals):
        row = dict(zip(key_columns, key, strict=True))
        row["合計枚数"] = totals[key]
        rows.append(row)
    return TableData("在庫集計", SUMMARY_COLUMNS, rows)


def build_pallet_list(pallets: list[Any], today: date) -> TableData:
    rows: list[dict[str, Any]] = []
    for pallet in pallets:
        if not isinstance(pallet, dict):
            continue
        items = pallet.get("items")
        item_list = items if isinstance(items, list) else []
        received_date = first_value(pallet, "received_date", "received_at")
        rows.append(
            {
                "パレット番号": as_text(first_value(pallet, "pallet_number", "pallet_no")),
                "保管場所": as_text(first_value(pallet, "location_code", "location")),
                "入庫日": as_text(received_date),
                "在庫日数": inventory_days(received_date, today),
                "色": as_text(first_value(pallet, "color_key", "color", "color_mode")),
                "更新日時": as_text(first_value(pallet, "updated_at")),
                "明細数": len(item_list),
                "合計枚数": sum(as_int(item.get("sheet_count")) for item in item_list if isinstance(item, dict)),
            }
        )
    return TableData("パレット一覧", PALLET_COLUMNS, rows)


def build_expanded_table(name: str, values: Any) -> TableData:
    rows: list[dict[str, Any]] = []
    source = values if isinstance(values, list) else []
    for value in source:
        if not isinstance(value, dict):
            rows.append({"値": value})
            continue

        base = {k: v for k, v in value.items() if k != "items"}
        flat_base = flatten_value(base)
        items = value.get("items")
        if isinstance(items, list) and items:
            for item in items:
                flat_item = flatten_value(item, "items") if isinstance(item, dict) else {"items": item}
                rows.append({**flat_base, **flat_item})
        else:
            rows.append(flat_base)

    return TableData(name, ordered_columns(rows, MAJOR_COLUMN_PRIORITY), rows)


def build_tables(data: Any, today: date | None = None) -> list[TableData]:
    if not isinstance(data, dict):
        raise ValueError("JSONの最上位はオブジェクトである必要があります。")
    today = today or date.today()
    pallets_value = data.get("pallets", [])
    pallets = pallets_value if isinstance(pallets_value, list) else []
    detail = build_inventory_detail(pallets, today)
    return [
        detail,
        build_inventory_summary(detail),
        build_pallet_list(pallets, today),
        build_expanded_table("出庫履歴", data.get("shipments", [])),
        build_expanded_table("マップメモ", data.get("map_notes", [])),
    ]


def safe_sheet_name(name: str, used_names: set[str]) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", name).strip()[:31] or "Sheet"
    candidate = cleaned
    index = 2
    while candidate in used_names:
        suffix = f"_{index}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        index += 1
    used_names.add(candidate)
    return candidate


def write_workbook(output_path: Path, tables: list[TableData]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_names: set[str] = set()

    for table in tables:
        sheet = workbook.create_sheet(title=safe_sheet_name(table.name, used_names))
        columns = table.columns or ["メッセージ"]
        rows = list(table.rows)
        if not table.columns:
            rows = [{"メッセージ": "データなし"}]

        sheet.append(columns)
        for row in rows:
            values = []
            for column in columns:
                value = row.get(column, "")
                if column in NUMERIC_COLUMNS:
                    values.append("" if value in (None, "") else as_int(value))
                elif column in TEXT_COLUMNS:
                    values.append(as_text(value))
                else:
                    values.append(as_text(value))
            sheet.append(values)

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="334155")

        for column_cells in sheet.columns:
            header = as_text(column_cells[0].value)
            max_len = max(len(as_text(cell.value)) for cell in column_cells)
            letter = get_column_letter(column_cells[0].column)
            sheet.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)
            if header in TEXT_COLUMNS:
                for cell in column_cells[1:]:
                    cell.number_format = "@"
            elif header in NUMERIC_COLUMNS:
                for cell in column_cells[1:]:
                    cell.number_format = "0"

    workbook.save(output_path)


class LocationGridGuide(QWidget):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setMinimumSize(620, 500)

    def grid_edges(self, start: int, length: int, count: int) -> list[int]:
        return [start + (i * length) // count for i in range(count + 1)]

    def draw_blocked_cell(self, painter: QPainter, rect: QRect) -> None:
        painter.fillRect(rect.adjusted(1, 1, -1, -1), QColor("#E5E7EB"))
        painter.setPen(QPen(QColor("#9CA3AF"), 1))
        step = 8
        start = rect.left() - rect.height()
        stop = rect.right() + rect.height()
        for x in range(start, stop, step):
            painter.drawLine(x, rect.bottom(), x + rect.height(), rect.top())

    def paintEvent(self, event) -> None:
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.fillRect(self.rect(), QColor("#F8FAFC"))

        title_rect = QRect(0, 8, self.width(), 28)
        painter.setPen(QColor("#111827"))
        painter.setFont(QFont("Yu Gothic UI", 13, QFont.Bold))
        painter.drawText(title_rect, Qt.AlignCenter, "位置確認用 真上ビュー")

        bounds = self.rect().adjusted(72, 58, -72, -74)
        painter.fillRect(bounds, QColor("#FFFFFF"))
        x_edges = self.grid_edges(bounds.left(), bounds.width(), GRID_COLUMNS)
        y_edges = self.grid_edges(bounds.top(), bounds.height(), GRID_ROWS)

        for col in range(GRID_COLUMNS):
            cell_rect = QRect(x_edges[col], bounds.top(), x_edges[col + 1] - x_edges[col], bounds.height())
            if is_aisle_column(col):
                painter.fillRect(cell_rect, QColor("#FEF3C7"))

        for row in range(GRID_ROWS):
            for col in range(GRID_COLUMNS):
                if is_blocked_grid_cell(col, row):
                    rect = QRect(
                        x_edges[col],
                        y_edges[row],
                        x_edges[col + 1] - x_edges[col],
                        y_edges[row + 1] - y_edges[row],
                    )
                    self.draw_blocked_cell(painter, rect)

        painter.setPen(QPen(QColor("#CBD5E1"), 1))
        for x in x_edges:
            painter.drawLine(x, bounds.top(), x, bounds.bottom())
        for y in y_edges:
            painter.drawLine(bounds.left(), y, bounds.right(), y)
        painter.setPen(QPen(QColor("#334155"), 2))
        painter.drawRect(bounds)

        painter.save()
        painter.setOpacity(0.48)
        painter.setPen(QColor("#475569"))
        painter.setFont(QFont("Yu Gothic UI", 18, QFont.Bold))
        painter.drawText(bounds, Qt.AlignCenter, "見取り図のみ　アイテム情報は表示しません。")
        painter.restore()

        for col in range(GRID_COLUMNS):
            x = (x_edges[col] + x_edges[col + 1]) // 2
            label = grid_column_display_label(col)
            aisle = is_aisle_column(col)
            painter.setPen(QColor("#92400E") if aisle else QColor("#1E3A8A"))
            painter.setFont(QFont("Yu Gothic UI", 10 if aisle else 17, QFont.Bold))
            top_rect = QRect(x - 44, bounds.top() - 34, 88, 28)
            bottom_rect = QRect(x - 44, bounds.bottom() + 6, 88, 28)
            painter.drawText(top_rect, Qt.AlignCenter, label)
            painter.drawText(bottom_rect, Qt.AlignCenter, label)

        painter.setPen(QColor("#1E3A8A"))
        painter.setFont(QFont("Yu Gothic UI", 13, QFont.Bold))
        for row in range(GRID_ROWS):
            y = (y_edges[row] + y_edges[row + 1]) // 2
            label = f"{row + 1:02d}"
            painter.drawText(QRect(bounds.left() - 58, y - 12, 50, 24), Qt.AlignRight | Qt.AlignVCenter, label)
            painter.drawText(QRect(bounds.right() + 8, y - 12, 50, 24), Qt.AlignLeft | Qt.AlignVCenter, label)

        entrance_x = (x_edges[4] + x_edges[6]) // 2
        painter.setPen(QColor("#DC2626"))
        painter.setFont(QFont("Yu Gothic UI", 20, QFont.Bold))
        painter.drawText(QRect(entrance_x - 88, bounds.bottom() + 32, 176, 34), Qt.AlignCenter, "－入口－")


class MainWindow(QMainWindow):
    def __init__(self, icon: QIcon | None = None) -> None:
        super().__init__()
        self.setWindowTitle("在庫データExcel変換ツール")
        if icon is not None and not icon.isNull():
            self.setWindowIcon(icon)
        self.resize(980, 620)
        self.setAcceptDrops(True)
        self.today = date.today()
        self.current_file: Path | None = None
        self.tables: list[TableData] = []

        self.select_button = QPushButton("JSONを選択")
        self.select_button.setIcon(self.style().standardIcon(QStyle.SP_DialogOpenButton))
        self.select_button.clicked.connect(self.choose_file)
        self.export_button = QPushButton("Excelに変換して保存")
        self.export_button.setIcon(self.style().standardIcon(QStyle.SP_DialogSaveButton))
        self.export_button.clicked.connect(self.export_excel)
        self.export_button.setEnabled(False)
        self.path_label = QLabel("inventory-data-*.json をドラッグ&ドロップ、またはJSONを選択してください")
        self.path_label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        top = QHBoxLayout()
        top.addWidget(self.select_button)
        top.addWidget(self.path_label, 1)
        top.addWidget(self.export_button)

        self.drop_label = QLabel("ここへ inventory-data-*.json をドラッグ&ドロップできます")
        self.drop_label.setAlignment(Qt.AlignCenter)
        self.drop_label.setFrameShape(QFrame.StyledPanel)
        self.drop_label.setStyleSheet(
            "QLabel { border: 1px dashed #64748B; color: #334155; padding: 12px; background: #F8FAFC; }"
        )

        self.grid_guide = LocationGridGuide()
        self.status_label = QLabel("JSONを読み込み、読み込んだ順番のままExcel形式に変換します。")
        self.status_label.setStyleSheet("QLabel { color: #334155; }")

        layout = QVBoxLayout()
        layout.addLayout(top)
        layout.addWidget(self.drop_label)
        layout.addWidget(self.grid_guide, 1)
        layout.addWidget(self.status_label)

        root = QWidget()
        root.setLayout(layout)
        self.setCentralWidget(root)

    def dragEnterEvent(self, event: QDragEnterEvent) -> None:
        if any(url.toLocalFile().lower().endswith(".json") for url in event.mimeData().urls()):
            event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent) -> None:
        for url in event.mimeData().urls():
            local_file = url.toLocalFile()
            if local_file.lower().endswith(".json"):
                self.load_file(Path(local_file))
                event.acceptProposedAction()
                return

    def choose_file(self) -> None:
        filename, _ = QFileDialog.getOpenFileName(
            self,
            "JSONファイルを選択",
            str(Path.home() / "Downloads"),
            "JSON files (*.json)",
        )
        if filename:
            self.load_file(Path(filename))

    def show_error(self, title: str, message: str) -> None:
        QMessageBox.critical(self, title, message)
        self.status_label.setText(message)

    def load_file(self, path: Path) -> None:
        try:
            text = path.read_text(encoding="utf-8")
            data = json.loads(text)
            self.tables = build_tables(data, self.today)
        except UnicodeDecodeError:
            self.show_error("読み込みエラー", "JSONファイルをUTF-8として読み込めませんでした。")
            return
        except json.JSONDecodeError as exc:
            self.show_error("JSON形式エラー", f"JSON形式が不正です。\n{exc}")
            return
        except Exception as exc:
            self.show_error("読み込みエラー", f"読み込み中にエラーが発生しました。\n{exc}")
            return

        self.current_file = path
        self.path_label.setText(str(path))
        self.export_button.setEnabled(True)
        total_rows = sum(len(table.rows) for table in self.tables)
        sheet_summary = " / ".join(f"{table.name}:{len(table.rows)}行" for table in self.tables)
        self.status_label.setText(f"{len(self.tables)}シート、合計{total_rows}行を読み込みました。{sheet_summary}")

    def export_excel(self) -> None:
        if self.current_file is None or not self.tables:
            return
        default_path = str(self.current_file.with_suffix(".xlsx"))
        output_path, _ = QFileDialog.getSaveFileName(self, "Excel出力", default_path, "Excel files (*.xlsx)")
        if not output_path:
            return
        output = Path(output_path if output_path.lower().endswith(".xlsx") else f"{output_path}.xlsx")

        try:
            write_workbook(output, self.tables)
        except Exception as exc:
            QMessageBox.critical(self, "Excel出力エラー", f"xlsx出力に失敗しました。\n{exc}")
            return
        self.status_label.setText(f"Excelファイルを保存しました: {output}")
        QMessageBox.information(self, "Excel出力完了", f"保存しました。\n{output}")


def main() -> int:
    set_windows_app_user_model_id()
    app = QApplication(sys.argv)
    icon = app_icon()
    if not icon.isNull():
        app.setWindowIcon(icon)
    window = MainWindow(icon)
    if len(sys.argv) > 1:
        window.load_file(Path(sys.argv[1]))
    window.show()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
